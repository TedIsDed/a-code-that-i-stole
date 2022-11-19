from openpyxl import load_workbook
import sqlite3

with open('students.db', 'w') as f:
    f.write('')

db = sqlite3.connect('students.db')
cursor = db.cursor()


students = load_workbook('Students.xlsx')
students_ws = students.active
col = 1
row = 1

while students_ws.cell(column=col, row=1).value is not None and students_ws.cell(column=col, row=1).value != '=""':
    col += 1

while students_ws.cell(column=1, row=row).value is not None and students_ws.cell(column=1, row=row).value != '=""':
    row += 1

rows = [[students_ws.cell(column=c, row=r).value for c in range(1, col)] for r in range(1, row)]

for i, v in enumerate(rows):
    if i == 0:
        stmt = 'CREATE TABLE IF NOT EXISTS excel('
        for j, w in enumerate(v):
            stmt += w.lower().replace(' ', '_')
            stmt += ' '
            sqlt_type = 'INT' if (ct := type(rows[i + 1][j])) == int else ('FLOAT' if ct == float else 'TEXT')
            stmt += sqlt_type
            stmt += ', '

        stmt += ')'
        stmt = list(stmt)
        stmt[-3:-1] = []
        stmt = ''.join(stmt)

        cursor.execute(stmt)

    else:
        stmt = 'INSERT INTO excel VALUES('
        for j, w in enumerate(v):
            stmt += str(w) if not isinstance(w, str) else f'"{w}"'
            stmt += ', '

        stmt += ')'
        stmt = list(stmt)
        stmt[-3:-1] = []
        stmt = ''.join(stmt)
        cursor.execute(stmt)

db.commit()
db.close()
